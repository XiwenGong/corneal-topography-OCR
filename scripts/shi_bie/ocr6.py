import os
import datetime
import openpyxl
from openpyxl.styles import Alignment, Font
from collections import defaultdict

def save_to_excel(results, classify_result):
    """
    保存识别结果为宽表结构，按 classify_result（图片名->类型/别名）分组，每组前插入合并单元格的类型行。
    :param results: {图片名: [ {area_name, type, coords, text}, ... ] }
    :param classify_result: {图片名: 图片类型/别名}
    """
    # 1. 确定保存路径
    current_dir = os.path.dirname(os.path.abspath(__file__))
    scripts_dir = os.path.dirname(current_dir)
    project_dir = os.path.dirname(scripts_dir)
    results_dir = os.path.join(project_dir, 'results')
    os.makedirs(results_dir, exist_ok=True)
    now_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    save_path = os.path.join(results_dir, f'ocr_results_{now_str}.xlsx')

    # 2. 用 classify_result 获取每张图片的类型
    img_type_map = {}  # {图片名: 图片类型}
    for img_name in results:
        img_type_map[img_name] = classify_result.get(img_name, '未知类型')

    # 3. 按类型分组
    type_group = defaultdict(list)  # {type_name: [图片名]}
    for img_name, type_name in img_type_map.items():
        type_group[type_name].append(img_name)

    # 4. 创建excel并写入表头
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '识别结果'
    headers = ['图片类型', '图片名', '识别区1', '识别区2', '识别区3', '识别区4']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 5. 按类型分组写入数据
    row_idx = 2  # 当前写入的行号
    for type_name, img_list in type_group.items():
        # 插入类型行，合并后面所有单元格
        ws.append([type_name] + [''] * (len(headers) - 1))
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=13)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
        row_idx += 1
        # 插入该类型下所有图片数据
        for img_name in img_list:
            area_list = results[img_name]
            region_contents = {1: [], 2: [], 3: [], 4: []}
            for area in area_list:
                region_type = area.get('type', '')
                text = area.get('text', '')
                if region_type in [1, 2, 3, 4]:
                    region_contents[region_type].append(text)
            row = [type_name, img_name]
            for i in range(1, 5):
                if region_contents[i]:
                    cell_text = '\n'.join([f'【{idx+1}】\n{t}' for idx, t in enumerate(region_contents[i])])
                else:
                    cell_text = ''
                row.append(cell_text)
            ws.append(row)
            row_idx += 1

    # 6. 设置所有数据单元格自动换行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 7. 优化列宽：每列宽度=该列所有单元格中最长一行（被\n分隔）长度+2
    for col in ws.columns:
        max_line_len = 10
        for cell in col:
            try:
                lines = str(cell.value).split('\n') if cell.value is not None else ['']
                for line in lines:
                    chinese_count = sum(1 for c in line if '\u4e00' <= c <= '\u9fff')
                    other_count = len(line) - chinese_count
                    line_len = other_count + chinese_count * 2
                    max_line_len = max(max_line_len, line_len)
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_line_len + 2
    # 8. 保存
    wb.save(save_path)
    print(f"[保存] 识别结果已保存到: {save_path}") 