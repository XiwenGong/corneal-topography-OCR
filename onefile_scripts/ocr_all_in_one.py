import sys
import os
import shutil
import pickle
import importlib.util
import numpy as np
import cv2
import base64
import urllib
import requests
import pytesseract
from PIL import Image
import datetime
import openpyxl
from openpyxl.styles import Alignment, Font
from collections import defaultdict
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QLabel, QPushButton, 
                            QVBoxLayout, QHBoxLayout, QProgressBar, QTextEdit, QMessageBox)
from PyQt6.QtGui import QPixmap, QPainter, QColor, QFont, QFontMetrics, QDragEnterEvent, QDropEvent
from PyQt6.QtCore import Qt, QRect, QSize, pyqtSignal, QMimeData, QThread

def get_project_root():
    """
    获取项目根目录
    如果是exe运行，则返回exe所在目录
    如果是脚本运行，则返回项目根目录
    """
    if getattr(sys, 'frozen', False):
        # 如果是exe运行
        return os.path.dirname(sys.executable)
    else:
        # 如果是脚本运行
        current_dir = os.path.dirname(os.path.abspath(__file__))  # onefile_scripts
        return os.path.dirname(current_dir)  # 项目根目录

def get_resource_path(relative_path):
    """
    获取资源文件的绝对路径
    如果是exe运行，则从临时目录获取
    如果是脚本运行，则从项目目录获取
    """
    if getattr(sys, 'frozen', False):
        # 如果是exe运行
        base_path = sys._MEIPASS
    else:
        # 如果是脚本运行
        base_path = get_project_root()
    return os.path.join(base_path, relative_path)

def get_mu_ban_dir():
    """
    获取mu_ban目录
    """
    mu_ban_dir = get_resource_path('mu_ban')
    os.makedirs(mu_ban_dir, exist_ok=True)
    return mu_ban_dir

def get_lin_shi_dir():
    """
    获取lin_shi目录
    """
    lin_shi_dir = get_resource_path('lin_shi')
    os.makedirs(lin_shi_dir, exist_ok=True)
    return lin_shi_dir

def get_results_dir():
    """
    获取results目录
    """
    results_dir = get_resource_path('results')
    os.makedirs(results_dir, exist_ok=True)
    return results_dir

# 原 ocr1.py 内容
class DropWindow(QMainWindow):
    """
    拖拽窗口类
    """
    def __init__(self):
        super().__init__()
        self.setWindowTitle('拖拽文件/文件夹到这里')
        self.setGeometry(100, 100, 400, 300)
        
        # 创建中央部件和布局
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        # 创建提示标签
        self.label = QLabel('将文件或文件夹拖拽到这里')
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.label)
        
        # 设置接受拖拽
        self.setAcceptDrops(True)
        
        # 确保目标文件夹存在
        self.target_dir = os.path.join(get_lin_shi_dir(), 'dai_shi_bie')
        print(f"目标文件夹路径: {self.target_dir}")  # 调试信息
        os.makedirs(self.target_dir, exist_ok=True)

    def dragEnterEvent(self, event: QDragEnterEvent):
        """
        处理拖拽进入事件
        """
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        """
        处理拖拽放下事件
        """
        # 获取拖拽的文件/文件夹路径
        urls = event.mimeData().urls()
        paths = [url.toLocalFile() for url in urls]
        
        # 收集所有图片文件
        image_files = []
        for path in paths:
            if os.path.isfile(path):
                # 如果是文件，检查是否为图片
                if self.is_image_file(path):
                    image_files.append(path)
            else:
                # 如果是文件夹，递归查找所有图片
                for root, _, files in os.walk(path):
                    for file in files:
                        file_path = os.path.join(root, file)
                        if self.is_image_file(file_path):
                            image_files.append(file_path)
        
        # 复制图片到目标文件夹
        if image_files:
            copied_files = []
            overwritten_files = []
            for image_file in image_files:
                try:
                    # 获取文件名
                    file_name = os.path.basename(image_file)
                    # 构建目标路径
                    target_path = os.path.join(self.target_dir, file_name)
                    print(f"正在复制: {image_file} -> {target_path}")  # 调试信息
                    
                    # 检查目标文件是否已存在
                    if os.path.exists(target_path):
                        overwritten_files.append(file_name)
                    
                    # 复制文件
                    shutil.copy2(image_file, target_path)
                    copied_files.append(file_name)
                except Exception as e:
                    print(f"复制文件失败: {e}")  # 调试信息
                    QMessageBox.warning(self, '错误', f'复制文件 {file_name} 失败: {str(e)}')
            
            if copied_files:
                # 构建成功消息
                success_msg = f'已成功复制 {len(copied_files)} 个图片文件到:\n{self.target_dir}'
                if overwritten_files:
                    success_msg += f'\n\n其中覆盖了 {len(overwritten_files)} 个已存在的文件:\n' + '\n'.join(overwritten_files)
                # 显示成功消息
                QMessageBox.information(self, '成功', success_msg)
                self.close()  # 拖拽成功后自动关闭窗口
            else:
                # 显示失败消息
                QMessageBox.warning(self, '失败', '没有成功复制任何文件')
        else:
            # 显示无图片消息
            QMessageBox.warning(self, '警告', '没有找到任何图片文件')

    def is_image_file(self, file_path: str) -> bool:
        """
        判断文件是否为图片
        
        @param file_path {str} 文件路径
        @return {bool} 是否为图片
        """
        # 获取文件扩展名（小写）
        ext = os.path.splitext(file_path)[1].lower()
        # 检查是否为常见图片格式
        return ext in ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp']

def run_ocr1():
    """
    运行拖拽窗口，供主流程调用
    """
    app = QApplication(sys.argv)
    window = DropWindow()
    window.show()
    app.exec()

# 原 ocr5.py 内容
class ProgressWindow(QWidget):
    def __init__(self, total_copy, total_judge, total_recognize):
        super().__init__()
        self.setWindowTitle('处理进度')
        self.setGeometry(400, 300, 400, 220)
        self.total_copy = total_copy
        self.total_judge = total_judge
        self.total_recognize = total_recognize
        self.copy_count = 0
        self.judge_count = 0
        self.recognize_count = 0
        layout = QVBoxLayout()
        self.text = QTextEdit()
        self.text.setReadOnly(True)
        self.text.setStyleSheet('background:#fafafa;')
        self.text.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        layout.addWidget(self.text)
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(max(total_copy, total_judge, total_recognize))
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)
        self.setLayout(layout)
        self.setWindowFlag(Qt.WindowType.WindowCloseButtonHint, False)
        self.update_text()

    def update_text(self):
        lines = []
        lines.append(f'已复制: {self.copy_count} / {self.total_copy}')
        if self.copy_count >= self.total_copy:
            lines.append(f'已判断: {self.judge_count} / {self.total_judge}')
        if self.judge_count >= self.total_judge and self.total_judge > 0:
            lines.append(f'已识别: {self.recognize_count} / {self.total_recognize}')
        self.text.setText('\n'.join(lines))

    def update_copy(self, val):
        self.copy_count = val
        self.progress_bar.setMaximum(self.total_copy)
        self.progress_bar.setValue(val)
        self.update_text()
        QApplication.processEvents()

    def update_judge(self, val):
        self.judge_count = val
        self.progress_bar.setMaximum(self.total_judge)
        self.progress_bar.setValue(val)
        self.update_text()
        QApplication.processEvents()

    def update_recognize(self, val):
        self.recognize_count = val
        self.progress_bar.setMaximum(self.total_recognize)
        self.progress_bar.setValue(val)
        self.update_text()
        QApplication.processEvents()

def show_progress_window(total_copy, total_judge, total_recognize, process_func):
    """
    显示进度窗口，process_func需接收3个回调：update_copy, update_judge, update_recognize
    """
    app = QApplication.instance() or QApplication(sys.argv)
    win = ProgressWindow(total_copy, total_judge, total_recognize)
    win.show()
    def update_copy(val):
        win.update_copy(val)
    def update_judge(val):
        win.update_judge(val)
    def update_recognize(val):
        win.update_recognize(val)
    process_func(update_copy, update_judge, update_recognize)
    win.close()

# 原 ocr2.py 内容
def load_judge_functions(pkl_path: str) -> dict:
    """
    从pkl文件中加载所有别名的判别函数
    只查找 pkl2 下的 图片类型判别方案 字段，且函数名必须为 judge
    """
    judge_functions = {}
    try:
        with open(pkl_path, 'rb') as f:
            data = pickle.load(f)
        for alias in data:
            if alias == '__global__':
                continue
            func_code = None
            # 只查找 pkl2 下的 图片类型判别方案
            if 'pkl2' in data[alias] and '图片类型判别方案' in data[alias]['pkl2']:
                func_code = data[alias]['pkl2']['图片类型判别方案']
            if func_code:
                try:
                    spec = importlib.util.spec_from_loader('temp_module', loader=None)
                    module = importlib.util.module_from_spec(spec)
                    # 注入依赖
                    module.__dict__['np'] = np
                    module.__dict__['cv2'] = cv2
                    exec(func_code, module.__dict__)
                    if hasattr(module, 'judge'):
                        judge_functions[alias] = getattr(module, 'judge')
                except Exception as e:
                    print(f"加载别名 {alias} 的判别函数失败: {e}")
    except Exception as e:
        print(f"读取pkl文件失败: {e}")
    return judge_functions

def judge_images(images_dir: str, judge_functions: dict) -> dict:
    """
    对图片进行类型判别
    
    @param images_dir {str} 图片目录
    @param judge_functions {Dict[str, Callable]} 判别函数字典
    @return {Dict[str, str]} 图片名到类型的映射
    """
    results = {}
    
    # 遍历所有图片
    for filename in os.listdir(images_dir):
        if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp')):
            image_path = os.path.join(images_dir, filename)
            
            # 读取图片为 np.ndarray
            image = cv2.imread(image_path)
            if image is None:
                print(f"无法读取图片: {filename}")
                results[filename] = 'unreadable'
                continue
            
            # 尝试每个判别函数
            image_type = 'unknown'
            for alias, judge_func in judge_functions.items():
                try:
                    if judge_func(image):
                        image_type = alias
                        break
                except Exception as e:
                    print(f"判别图片 {filename} 时出错: {e}")
                    continue
            
            results[filename] = image_type
            
    return results

def classify_with_progress():
    # 统计待分类图片
    images_dir = os.path.join(get_lin_shi_dir(), 'dai_shi_bie')
    image_files = [f for f in os.listdir(images_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp'))]
    total = len(image_files)
    classify_result = {}
    def process_func(update_copy, update_judge, update_recognize):
        # 复制阶段（已完成，直接更新）
        update_copy(total)
        # 判断阶段
        pkl_path = os.path.join(get_mu_ban_dir(), 'shared_data.pkl')
        judge_functions = load_judge_functions(pkl_path)
        if not judge_functions:
            print("警告: 没有找到任何判别函数")
            return
        result = judge_images(images_dir, judge_functions)
        for i, _ in enumerate(result):
            update_judge(i+1)
        # 识别阶段跳过
        update_recognize(0)
        classify_result.update(result)
    show_progress_window(total, total, total, process_func)
    return classify_result

# 原 ocr3.py 内容
def read_baidu_ocr_key():
    key_path = os.path.join(get_mu_ban_dir(), 'baidu_ocr_key.txt')
    api_key = ''
    secret_key = ''
    try:
        with open(key_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line.startswith('API_KEY'):
                    # 解析等号右侧内容，去除引号和空格
                    api_key = line.split('=', 1)[1].strip().strip('"').strip("'")
                elif line.startswith('SECRET_KEY'):
                    secret_key = line.split('=', 1)[1].strip().strip('"').strip("'")
        return api_key, secret_key
    except Exception as e:
        print(f"读取百度OCR密钥文件失败: {e}")
        return '', ''

def get_access_token():
    API_KEY, SECRET_KEY = read_baidu_ocr_key()
    if not API_KEY or not SECRET_KEY:
        print("API_KEY 或 SECRET_KEY 为空，请在 mu_ban/baidu_ocr_key.txt 中填写！")
        return None
    try:
        url = f'https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id={API_KEY}&client_secret={SECRET_KEY}'
        resp = requests.get(url)
        data = resp.json()
        return data.get('access_token', None)
    except Exception as e:
        print(f"获取access_token失败: {e}")
        return None

def get_file_content_as_base64(image_array, urlencoded=False):
    _, img_encoded = cv2.imencode('.png', image_array)
    content = base64.b64encode(img_encoded.tobytes()).decode("utf8")
    if urlencoded:
        content = urllib.parse.quote_plus(content)
    return content

def baidu_ocr(image_array, access_token):
    try:
        url = f"https://aip.baidubce.com/rest/2.0/ocr/v1/accurate_basic?access_token={access_token}"
        image_base64 = get_file_content_as_base64(image_array, True)
        payload = f'image={image_base64}&language_type=CHN_ENG&detect_direction=false&paragraph=false&probability=false&multidirectional_recognize=false'
        headers = {
            'Content-Type': 'application/x-www-form-urlencoded',
            'Accept': 'application/json'
        }
        response = requests.request("POST", url, headers=headers, data=payload.encode("utf-8"))
        result = response.json()
        if "words_result" in result:
            words = result["words_result"]
            formatted_text = ""
            for word in words:
                formatted_text += word["words"] + "\n"
            return formatted_text.strip()
        else:
            return ""
    except Exception as e:
        print(f"百度OCR识别出错: {str(e)}")
        return ""

def tesseract_ocr(image_array, access_token=None):
    # image_array: numpy.ndarray (BGR)
    try:
        # 转为RGB
        rgb_img = cv2.cvtColor(image_array, cv2.COLOR_BGR2RGB)
        pil_img = Image.fromarray(rgb_img)
        text = pytesseract.image_to_string(pil_img, lang='chi_sim+eng')
        return text.strip()
    except Exception as e:
        print(f"Tesseract OCR识别出错: {e}")
        return ""

OCR_ENGINES = {
    'tesseractOCR': tesseract_ocr,
    '百度OCR': baidu_ocr
}

def load_recognition_schemes(pkl_path):
    """
    读取shared_data.pkl，获取每个类别的识别区划分方案、预处理、后处理、OCR引擎
    返回：
      schemes: {类别: {区名: {pre, post, ocr_engine, 区域坐标}}}
      global_basic_types: {basic_type_x: {...}}
    """
    with open(pkl_path, 'rb') as f:
        data = pickle.load(f)
    schemes = {}
    global_basic_types = {}
    for alias in data:
        if alias == '__global__':
            # 提取全局基础类型
            for k, v in data[alias].items():
                if k.startswith('basic_type_'):
                    global_basic_types[k] = v
            continue
        # 识别区划分方案、预处理、后处理、OCR引擎
        if 'pkl5' in data[alias]:
            area_info = data[alias]['pkl5']  # 假设结构为 {区名: {...}}
            schemes[alias] = area_info
    return schemes, global_basic_types

def exec_code(code_str, local_vars):
    """
    动态执行代码字符串，local_vars为本地变量字典
    """
    try:
        exec(code_str, {}, local_vars)
    except Exception as e:
        print(f'执行动态代码出错: {e}')

class RecognizeThread(QThread):
    progress_signal = pyqtSignal(int)  # 当前已识别数量
    result_signal = pyqtSignal(dict)   # 最终识别结果

    def __init__(self, classify_result):
        super().__init__()
        self.classify_result = classify_result

    def run(self):
        import numpy as np
        import cv2
        results = {}
        images_dir = os.path.join(get_lin_shi_dir(), 'dai_shi_bie')
        pkl_path = os.path.join(get_mu_ban_dir(), 'shared_data.pkl')
        schemes, global_basic_types = load_recognition_schemes(pkl_path)
        access_token = get_access_token()
        with open(pkl_path, 'rb') as f:
            all_data = pickle.load(f)
        images = list(self.classify_result.items())
        total = len(images)
        for idx, (img_name, img_type) in enumerate(images):
            img_path = os.path.join(images_dir, img_name)
            image = cv2.imread(img_path)
            if image is None:
                results[img_name] = [{'error': '图片无法读取'}]
                self.progress_signal.emit(idx + 1)
                continue
            img_result = []
            # 读取该图片的pkl3标注
            pkl3_data = all_data.get(img_type, {}).get('pkl3', {})
            boxes = pkl3_data.get('boxes', [])
            for box in boxes:
                box_type = box.get('type')
                area_name = box.get('area_name') or f'basic_type_{box_type}'
                pt1 = box.get('pt1')
                pt2 = box.get('pt2')
                if not (1 <= box_type <= 4):
                    continue
                basic_type_key = f'basic_type_{box_type}'
                area_info = global_basic_types.get(basic_type_key, {})
                if not pt1 or not pt2:
                    img_result.append({'area_name': area_name, 'type': box_type, 'coords': None, 'text': '无区域坐标'})
                    continue
                x1, y1 = pt1
                x2, y2 = pt2
                roi = image[min(y1, y2):max(y1, y2), min(x1, x2):max(x1, x2)]
                pre_code = area_info.get('预处理方案', '')
                if pre_code:
                    local_vars = {'img': roi, 'np': np, 'cv2': cv2}
                    exec_code(pre_code, local_vars)
                    roi = local_vars.get('img', roi)
                ocr_engine_name = area_info.get('OCR引擎', 'tesseractOCR')
                ocr_func = OCR_ENGINES.get(ocr_engine_name, tesseract_ocr)
                if ocr_engine_name == '百度OCR':
                    ocr_result = ocr_func(roi, access_token)
                else:
                    ocr_result = ocr_func(roi)
                post_code = area_info.get('后处理方案', '')
                if post_code:
                    local_vars = {'text': ocr_result}
                    exec_code(post_code, local_vars)
                    ocr_result = local_vars.get('text', ocr_result)
                img_result.append({
                    'area_name': area_name,
                    'type': box_type,
                    'coords': (x1, y1, x2, y2),
                    'text': ocr_result
                })
            # 处理pkl5下的所有非basic_type_x区（如有）
            type_scheme = schemes.get(img_type)
            if type_scheme:
                for area_name, area_info in type_scheme.items():
                    if area_name.startswith('basic_type_'):
                        continue
                    coords = area_info.get('coords')
                    if not coords:
                        img_result.append({'area_name': area_name, 'type': None, 'coords': None, 'text': '无区域坐标'})
                        continue
                    x1, y1, x2, y2 = coords
                    roi = image[y1:y2, x1:x2]
                    pre_code = area_info.get('预处理方案', '')
                    if pre_code:
                        local_vars = {'img': roi, 'np': np, 'cv2': cv2}
                        exec_code(pre_code, local_vars)
                        roi = local_vars.get('img', roi)
                    ocr_engine_name = area_info.get('OCR引擎', 'tesseractOCR')
                    ocr_func = OCR_ENGINES.get(ocr_engine_name, tesseract_ocr)
                    if ocr_engine_name == '百度OCR':
                        ocr_result = ocr_func(roi, access_token)
                    else:
                        ocr_result = ocr_func(roi)
                    post_code = area_info.get('后处理方案', '')
                    if post_code:
                        local_vars = {'text': ocr_result}
                        exec_code(post_code, local_vars)
                        ocr_result = local_vars.get('text', ocr_result)
                    img_result.append({
                        'area_name': area_name,
                        'type': None,
                        'coords': (x1, y1, x2, y2),
                        'text': ocr_result
                    })
            results[img_name] = img_result
            self.progress_signal.emit(idx + 1)
        self.result_signal.emit(results)

def recognize_with_progress(classify_result):
    images_dir = os.path.join(get_lin_shi_dir(), 'dai_shi_bie')
    image_files = [f for f in os.listdir(images_dir) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp'))]
    total = len(image_files)
    ocr_result = {}

    app = QApplication.instance() or QApplication(sys.argv)
    win = ProgressWindow(total, total, total)
    win.show()
    win.update_copy(total)
    win.update_judge(total)

    # 启动识别线程
    thread = RecognizeThread(classify_result)
    def on_progress(val):
        win.update_recognize(val)
    def on_result(result):
        ocr_result.update(result)
        win.close()
        app.quit()
    thread.progress_signal.connect(on_progress)
    thread.result_signal.connect(on_result)
    thread.start()
    app.exec()
    return ocr_result

# 原 ocr4.py 内容
class ImageWindow(QWidget):
    def __init__(self, image_paths, on_index_change, save_callback=None):
        super().__init__()
        self.setWindowTitle('图片浏览')
        self.setGeometry(100, 100, 800, 600)
        self.image_paths = image_paths
        self.index = 0
        self.on_index_change = on_index_change
        self.save_callback = save_callback

        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.index_label = QLabel()
        self.index_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignBottom)
        self.index_label.setStyleSheet('font-size: 14px; color: #666;')

        prev_btn = QPushButton('上一张')
        next_btn = QPushButton('下一张')
        save_btn = QPushButton('保存')
        prev_btn.clicked.connect(self.prev_image)
        next_btn.clicked.connect(self.next_image)
        save_btn.clicked.connect(self.on_save_clicked)

        btn_layout = QHBoxLayout()
        btn_layout.addWidget(prev_btn)
        btn_layout.addWidget(next_btn)
        btn_layout.addWidget(save_btn)
        btn_layout.addStretch(1)
        btn_layout.addWidget(self.index_label)

        layout = QVBoxLayout()
        layout.addWidget(self.image_label, stretch=1)
        layout.addLayout(btn_layout)
        self.setLayout(layout)

        self.update_image()

    def update_image(self):
        img_path = self.image_paths[self.index]
        pixmap = QPixmap(img_path)
        if not pixmap.isNull():
            scaled = pixmap.scaled(self.image_label.size(), Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            self.image_label.setPixmap(scaled)
        else:
            self.image_label.setText('图片无法加载')
        self.index_label.setText(f'{self.index+1}/{len(self.image_paths)}')
        self.on_index_change(self.index, self.get_image_display_info())

    def get_image_display_info(self):
        img_path = self.image_paths[self.index]
        pixmap = QPixmap(img_path)
        label_size = self.image_label.size()
        if pixmap.isNull() or label_size.width() == 0 or label_size.height() == 0:
            return None
        img_w, img_h = pixmap.width(), pixmap.height()
        label_w, label_h = label_size.width(), label_size.height()
        scale = min(label_w / img_w, label_h / img_h)
        show_w, show_h = int(img_w * scale), int(img_h * scale)
        offset_x = (label_w - show_w) // 2
        offset_y = (label_h - show_h) // 2
        return {
            'img_w': img_w, 'img_h': img_h,
            'show_w': show_w, 'show_h': show_h,
            'offset_x': offset_x, 'offset_y': offset_y,
            'scale': scale
        }

    def resizeEvent(self, event):
        self.update_image()
        super().resizeEvent(event)

    def prev_image(self):
        if self.index > 0:
            self.index -= 1
            self.update_image()

    def next_image(self):
        if self.index < len(self.image_paths) - 1:
            self.index += 1
            self.update_image()

    def on_save_clicked(self):
        if self.save_callback:
            self.save_callback()

class ResultOverlayWindow(QWidget):
    save_requested = pyqtSignal(dict)  # 新增信号
    def __init__(self, image_paths, ocr_results, get_display_info_func, save_callback=None, classify_result=None):
        super().__init__()
        self.setWindowTitle('识别结果')
        self.setGeometry(950, 100, 800, 600)
        self.image_paths = image_paths
        self.ocr_results = ocr_results
        self.index = 0
        self.display_info = None
        self.get_display_info_func = get_display_info_func
        self.setMinimumSize(400, 300)
        self.setFont(QFont('Consolas', 12))
        self.setLayout(QVBoxLayout(self))
        self.save_callback = save_callback
        self.classify_result = classify_result

    def show_result(self, index, display_info):
        self.index = index
        self.display_info = display_info
        if display_info:
            self.resize(display_info['show_w'], display_info['show_h'])
        self.update()

    def paintEvent(self, event):
        super().paintEvent(event)
        if not self.display_info:
            return
        img_name = os.path.basename(self.image_paths[self.index])
        result = self.ocr_results.get(img_name, [])
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        if isinstance(result, list):
            for area_info in result:
                coords = area_info.get('coords')
                text = area_info.get('text', '')
                area_type = area_info.get('type')
                area_name = area_info.get('area_name')
                if coords:
                    x1, y1, x2, y2 = coords
                    # 保证左上-右下顺序
                    x1, x2 = sorted([x1, x2])
                    y1, y2 = sorted([y1, y2])
                    scale = self.display_info['scale']
                    offset_x = self.display_info['offset_x']
                    offset_y = self.display_info['offset_y']
                    rx1 = int(x1 * scale + offset_x)
                    ry1 = int(y1 * scale + offset_y)
                    rx2 = int(x2 * scale + offset_x)
                    ry2 = int(y2 * scale + offset_y)
                    rect = QRect(rx1, ry1, rx2 - rx1, ry2 - ry1)
                    print(f"[调试] 显示区: type={area_type}, area_name={area_name}, coords={coords}, rect=({rx1},{ry1},{rx2},{ry2}), text={text}")
                    painter.setPen(QColor(0, 0, 0))
                    lines = str(text).splitlines() if text else ['']
                    n_lines = len(lines)
                    if n_lines == 0:
                        continue
                    max_font_size = 32
                    area_height = max(ry2 - ry1, 1)
                    area_width = max(rx2 - rx1, 1)
                    font_size_h = int(area_height / n_lines * 0.8)
                    font_size = min(max_font_size, font_size_h)
                    font = QFont('Consolas', font_size)
                    metrics = QFontMetrics(font)
                    for try_size in range(font_size, 0, -1):
                        font.setPointSize(try_size)
                        metrics = QFontMetrics(font)
                        too_wide = False
                        for line in lines:
                            if metrics.horizontalAdvance(line) > area_width:
                                too_wide = True
                                break
                        if not too_wide:
                            font_size = try_size
                            break
                    painter.setFont(QFont('Consolas', font_size))
                    line_height = area_height / n_lines
                    for i, line in enumerate(lines):
                        line_rect = QRect(rx1, int(ry1 + i * line_height), area_width, int(line_height))
                        painter.drawText(line_rect, Qt.AlignmentFlag.AlignCenter, line)
                else:
                    print(f"[调试] 跳过显示: type={area_type}, area_name={area_name}, coords=None, text={text}")
        painter.end()

def show_image_and_results(ocr_results: dict, classify_result: dict, save_callback=None):
    images_dir = os.path.join(get_lin_shi_dir(), 'dai_shi_bie')
    image_paths = [os.path.join(images_dir, f) for f in os.listdir(images_dir)
                   if f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.webp'))]
    image_paths.sort()
    if not image_paths:
        print('未找到图片')
        return
    app = QApplication.instance() or QApplication(sys.argv)
    result_win = None
    def on_index_change(idx, display_info):
        nonlocal result_win
        if result_win is not None:
            result_win.show_result(idx, display_info)
            if display_info:
                result_win.resize(display_info['show_w'], display_info['show_h'])
        else:
            result_win = ResultOverlayWindow(image_paths, ocr_results, None, save_callback, classify_result)
            result_win.show_result(idx, display_info)
            result_win.show()
    def save_callback_wrapper():
        if save_callback:
            save_callback(ocr_results, classify_result)
    img_win = ImageWindow(image_paths, on_index_change, save_callback=save_callback_wrapper)
    img_win.show()
    app.exec()

# 原 ocr6.py 内容
def save_to_excel(results, classify_result):
    """
    保存识别结果为宽表结构，按 classify_result（图片名->类型/别名）分组，每组前插入合并单元格的类型行。
    :param results: {图片名: [ {area_name, type, coords, text}, ... ] }
    :param classify_result: {图片名: 图片类型/别名}
    """
    # 1. 确定保存路径
    results_dir = get_results_dir()
    os.makedirs(results_dir, exist_ok=True)
    now_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    save_path = os.path.join(results_dir, f'ocr_results_{now_str}.xlsx')

    # 2. 用 classify_result 获取每张图片的类型
    img_type_map = {}  # {图片名: 图片类型}
    for img_name in results:
        img_type_map[img_name] = classify_result.get(img_name, '未知类型')

    # 3. 按类型分组
    type_group = defaultdict(list)  # {type_name: [图片名]}
    for img_name, type_name in img_type_map.items():
        type_group[type_name].append(img_name)

    # 4. 创建excel并写入表头
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '识别结果'
    headers = ['图片类型', '图片名', '识别区1', '识别区2', '识别区3', '识别区4']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 5. 按类型分组写入数据
    row_idx = 2  # 当前写入的行号
    for type_name, img_list in type_group.items():
        # 插入类型行，合并后面所有单元格
        ws.append([type_name] + [''] * (len(headers) - 1))
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=13)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
        row_idx += 1
        # 插入该类型下所有图片数据
        for img_name in img_list:
            area_list = results[img_name]
            region_contents = {1: [], 2: [], 3: [], 4: []}
            for area in area_list:
                region_type = area.get('type', '')
                text = area.get('text', '')
                if region_type in [1, 2, 3, 4]:
                    region_contents[region_type].append(text)
            row = [type_name, img_name]
            for i in range(1, 5):
                if region_contents[i]:
                    cell_text = '\n'.join([f'【{idx+1}】\n{t}' for idx, t in enumerate(region_contents[i])])
                else:
                    cell_text = ''
                row.append(cell_text)
            ws.append(row)
            row_idx += 1

    # 6. 设置所有数据单元格自动换行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 7. 优化列宽：每列宽度=该列所有单元格中最长一行（被\n分隔）长度+2
    for col in ws.columns:
        max_line_len = 10
        for cell in col:
            try:
                lines = str(cell.value).split('\n') if cell.value is not None else ['']
                for line in lines:
                    chinese_count = sum(1 for c in line if '\u4e00' <= c <= '\u9fff')
                    other_count = len(line) - chinese_count
                    line_len = other_count + chinese_count * 2
                    max_line_len = max(max_line_len, line_len)
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_line_len + 2
    # 8. 保存
    wb.save(save_path)
    print(f"[保存] 识别结果已保存到: {save_path}")

def main():
    run_ocr1()
    classify_result = classify_with_progress()
    if not classify_result:
        print("分类失败，程序结束")
        return
    ocr_result = recognize_with_progress(classify_result)
    if not ocr_result:
        print("识别失败，程序结束")
        return

    # 保存回调，调用ocr6保存excel
    def save_callback(results, classify_result):
        print("[保存回调] 收到保存请求，正在保存到Excel...")
        save_to_excel(results, classify_result)

    show_image_and_results(ocr_result, classify_result, save_callback=save_callback)

    # 程序结束前清空临时文件夹
    temp_dir = os.path.join(get_lin_shi_dir(), 'dai_shi_bie')
    if os.path.exists(temp_dir):
        for f in os.listdir(temp_dir):
            fp = os.path.join(temp_dir, f)
            try:
                if os.path.isfile(fp) or os.path.islink(fp):
                    os.unlink(fp)
                elif os.path.isdir(fp):
                    shutil.rmtree(fp)
            except Exception as e:
                print(f'清理临时文件失败: {fp}, {e}')

if __name__ == '__main__':
    main() 